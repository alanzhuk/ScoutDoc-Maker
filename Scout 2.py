#writes to an excell document
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
#reads an excell document
import xlrd
import csv
#scrapes a webpage
from bs4 import BeautifulSoup as soup #parses HTML
from urllib.request import urlopen as uReq #opens webpage
#checks how simular different letters are
import Levenshtein
#Makes GUI
from tkinter import *
from tkinter import filedialog


master = Tk() 
T2 = Text(master, height=10, width=100, yscrollcommand=True) 



schools = []
names = []
local = []
team_code = []
aff_link = []
neg_link = []
URLs = []
cache = []
my_url = "https://hspolicy.debatecoaches.org/"
wb = Workbook()
ws = wb.active
file_path = ""
save_path = ""
dis_max = []
#opening connetion, grabbing the page for the main page to save proccessing powed
uClient = uReq(my_url)
page_html = uClient.read()
uClient.close()
page_soup_school = soup(page_html, "html.parser")

school_cache = ""
page_soup_team = soup(page_html, "html.parser")
schoolURL = page_soup_school.findAll("span",{"class":"wikilink"})
def extract_school(string): #finds the schools name from the html code
    string = str(string)
    tog = True
    start = 0
    while tog:
        if string[start] == ">":
            start += 1
            tog = False
        else:
            start += 1
    tog = True
    end = start
    while tog:
        if string[end] == "(":
            tog = False
        else:
            end += 1
    return (string[start:end].strip())
def seperate(code): #seperates partner names to find the team link. Code is the names, like ("Abugosh & Zhukovsky")
    name1 = ""
    name2 = ""
    lnames = []
    tog = True
    for i in code:
        if i != "&" and tog:
            name1 += i
        elif i != "&" and tog == False:
            name2 += i
        else:
            tog = False
    lnames.append(name1.strip())
    if name2 != "":
        lnames.append(name2.strip())
    return(lnames)
def report_seperate_neg(string): #takes the round reports and makes them digestible.
    start = 0
    progress = 0
    e = 0
    for i in string:
        if i == "1" and progress == 0:
            progress = 1
            start = e
        elif i == "n" and progress == 1:
            progress = 2
        elif i == "c" and progress == 2:
            break
        else:
            progress = 0
        e += 1
    progress = 0
    e = 0
    end = 0
    for i in string:
        if i == "2" and progress == 0:
            progress = 1
            end = e
        elif i == "n" and progress == 1:
            progress = 2
        elif i == "r" and progress == 2:
            break
        else:
            progress = 0
        e += 1
    return(string[start:end])
def report_seperate_aff(string): #takes the round reports and makes them digestible.
    start = 0
    progress = 0
    e = 0
    for i in string:
        if i == "1" and progress == 0:
            progress = 1
            start = e
        elif i == "a" and progress == 1:
            progress = 2
        elif i == "c" and progress == 2:
            break
        else:
            progress = 0
        e += 1
    progress = 0
    e = 0
    end = 0
    for i in string:
        if i == "1" and progress == 0:
            progress = 1
            end = e
        elif i == "n" and progress == 1:
            progress = 2
        elif i == "c" and progress == 2:
            break
        else:
            progress = 0
        e += 1
    return(string[start:end])
def check_duplicates(lst): #checks a list to see if there are duplicates. lst is the list.
    count = 0
    count2 = 0
    lst2 = lst
    for i in lst2:
        count2 = 0
        for a in lst2:
            if a == i and count != count2:
                lst.remove(i)
                break
                
            count2 += 1
        count += 1
    return(lst)
#puts all the schools and names into a list
def collect_schools(loc): #loc is the file location
    with open(loc, mode='r') as csv_file:
        readCSV = csv.reader(csv_file, delimiter=',')
        for row in readCSV: 
            schools.append(row[0])
            local.append(row[1])
            names.append(row[2])
            team_code.append(row[3])
        del schools[0]
        del local[0]
        del names[0]
        del team_code[0]
#returns the url for the school to use
def pick_school(school): #school is the name of the school in the schools array
    #grabs each school
    global schoolURL
    global cache
    global dis_max
    dis_max = []
    dic = { #dictionary to store schools and leventenshtein distances

        }
    storage = 6
    spacer = 0
    sort = []
    cache = []
    for i in schoolURL:
        if spacer > 5:
            dic[i] = int(Levenshtein.distance(school,extract_school(i.a)))
        spacer += 1
    sort = sorted(dic.items(), key=lambda x: x[1])
    for i in sort:
        print(Levenshtein.distance(school,str(i[0].a["href"])))
        dis_max.append(Levenshtein.distance(school,str(i[0].a["href"])))
        cache.append("https://hspolicy.debatecoaches.org/" + i[0].a["href"])
   #6 - 434
def pick_team(name,school,pos):
    pick_school(school)
    global school_cache
    global page_soup_team
    global cache
    global dis_max
    z = 0
    print(school)
    while z != len(cache) and z < 3 and name != "Names TBA":
        print(dis_max[z],"      ",z)
        if school == "Boston Latin Academy":
            print(str(cache[z]))
        team_url = str(cache[z])
        #opening connetion, grabbing the page
        uClient = uReq(team_url)
        page_html = uClient.read()
        uClient.close()
        page_soup_team = soup(page_html, "html.parser")
        school_cache = school
        #grabs each link
        teams = page_soup_team.findAll("span",{"class":"wikilink"})
        check_names = seperate(name)
        
        if len(check_names) == 2:
            for i in teams:
                if check_names[0] in str(i.a.text):
                    if check_names[1] in str(i.a.text):
                        if pos in str(i.a):
                            if pos == "Aff":
                                print(1)
                                aff_link.append("https://hspolicy.debatecoaches.org/" + i.a["href"])
                            else:
                                print(2)
                                neg_link.append("https://hspolicy.debatecoaches.org/" + i.a["href"])
                            return("https://hspolicy.debatecoaches.org/" + i.a["href"])
        elif len(check_names) == 1:
            for i in teams:
                if check_names[0] in str(i.a):
                    if pos in str(i.a):
                        if pos == "Aff":
                            print(1)
                            aff_link.append("https://hspolicy.debatecoaches.org/" + i.a["href"])
                        else:
                            print(2)
                            neg_link.append("https://hspolicy.debatecoaches.org/" + i.a["href"])
                        return("https://hspolicy.debatecoaches.org/" + i.a["href"])
        if dis_max[0] < 10:
            print("break")
            break
        z += 1
    print(3)
    if pos == "Aff":
        aff_link.append("none")
    else:
        neg_link.append("none")
def scrub(URL):
    my_url = str(URL)
    if my_url == "None":
        return "none"
    #opening connetion, grabbing the page
    uClient = uReq(my_url)
    page_html = uClient.read()
    uClient.close()
    page_soup = soup(page_html, "html.parser")
    #grabs each link
    cites = page_soup.findAll("h4",{"class":"title closed wikigeneratedheader"})
    reports = page_soup.findAll("div",{"name":"report"})
    export = ""
    if len(cites) != 0:
        for i in cites:
            export = export + "\n" + i.span.text
        return(export)
    else:
        if "Aff" in URL:
            temp = []
            for i in reports:
                temp.append(report_seperate_aff(i.p.text))
            temp = check_duplicates(temp)
            for i in temp:
                export = export + "\n" + i
            return(export)
        else:
            temp = []
            for i in reports:
                temp.append(report_seperate_neg(i.p.text))
            
            temp = check_duplicates(temp)
            for i in temp:
                export = export + "\n" + i
        return export
def gather(order,aff_args,neg_args): #compiles all the team information for appending into the excell document
    temp = []
    temp.append(schools[order])
    temp.append(local[order])
    temp.append(names[order])
    temp.append(team_code[order])
    temp.append(aff_link[order])
    temp.append(neg_link[order])
    temp.append(aff_args)
    temp.append(neg_args)
    return(temp)


#*actual script#*
def play(in_name,out_name):
    collect_schools(str(in_name))
    ws.append(["School Name","Location","Team Names","Team Code","Link to Aff Wiki","Link to Neg Wiki","Aff Arguments","Neg Arguments"])
    count3 = 0
    while count3 < len(schools):
        T2.insert(END," " + str(schools[count3]) + " - " + str(names[count3]))
        T2.insert(END,"\n ")
        master.update()
        ws.append(gather(count3,scrub(pick_team(names[count3],schools[count3],"Aff")),scrub(pick_team(names[count3],schools[count3],"Neg"))))
        count3 += 1
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                 dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    for row in ws.iter_rows():
        for cell in row:      
            cell.alignment =  cell.alignment.copy(wrapText=True)
        
    wb.save(str(out_name)+".xlsx")
    exit()
def choose_file():
    global file_path
    file_path = filedialog.askopenfilename()
    T3.delete('1.0', END)
    T3.insert(END,file_path)
    master.update()
def save_file():
    global save_path
    save_path = filedialog.asksaveasfilename()
    T4.delete('1.0', END)
    T4.insert(END,str(save_path)+".xlsx")
master.title("Scoutdoc Maker")
T = Text(master, height=3, width=100)
T3 = Text(master, height=1, width=100)
T4 = Text(master, height=1, width=100)
T.pack() 
T.insert(END, 'All you need to do is select your file(you can get a csv from tabroom under entries) and the save location(please type a name)       -       patent pending pending - Alan Zhukovsky') 
e1 = Button(master, text='choose file', width=25,command=lambda: choose_file())  
e2 = Button(master, text='choose save', width=25,command=lambda: save_file()) 
e1.pack()
T3.pack()
e2.pack()
T4.pack()
button = Button(master, text='run', width=25,command=lambda: play(file_path,save_path)) 
button.pack()
T2.pack()
mainloop()

